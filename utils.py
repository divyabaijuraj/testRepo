from openpyxl import load_workbook


class Utilities:

    def read_data_from_excel(file_name, sheet_name):
        data_list = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet_name]
        row_c = sh.max_row
        col_c = sh.max_column

        for i in range(2, row_c+1):
            row = []
            print(i,"rows")
            for j in range(1, col_c+1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)
            #print(j,"******jjjj")
            #print(sh.cell(row=i, column=j).value)
        print(data_list,"data list in excel")

        return data_list

    def writeData(file_name,sheet_name,data):
        data_list = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet_name]
        row_c = sh.max_row
        col_c = sh.max_column
        j=col_c
        for i in range(2, row_c + 1):
            row = []
            print(i, "rows")
            sh.cell(row=i, column=j).value=data
            print(i,j)
            wb.save(file_name)